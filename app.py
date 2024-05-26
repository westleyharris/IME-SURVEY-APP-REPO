from flask import Flask, request, redirect, url_for, render_template, flash
import base64
import requests
import io
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

# OpenAI API Key
api_key = os.getenv("OPENAI_API_KEY")

app = Flask(__name__)
app.secret_key = 'your_secret_key'

def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

def process_image_with_gpt4o(image_path):
    base64_image = encode_image(image_path)

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    payload = {
        "model": "gpt-4o",
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "Extract the motor plate information including RPM, NDE/ODE Bearing, DE Bearing, Horsepower, and thermal class."
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 300
    }

    response = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload)

    if response.status_code == 200:
        response_data = response.json()
        return response_data['choices'][0]['message']['content']
    else:
        print(f"Request failed: {response.status_code}")
        print(response.text)
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit_motor_info', methods=['POST'])
def submit_motor_info():
    area = request.form['area']
    functional_location = request.form['functional_location']
    equipment = request.form['equipment']
    assembly = request.form['assembly']
    component = request.form['component']
    plate_picture = request.files['plate_picture']
    component_picture = request.files['component_picture']

    # Save the uploaded plate picture
    plate_image_path = 'uploaded_plate.jpeg'
    plate_picture.save(plate_image_path)

    # Process the image with GPT-4o
    raw_text = process_image_with_gpt4o(plate_image_path)
    structured_info = {}
    if raw_text:
        for line in raw_text.split('\n'):
            if ':' in line:
                key, value = line.split(':', 1)
                key = key.strip('- *')  # Strip the unwanted characters
                value = value.strip('* ')  # Strip the unwanted characters
                structured_info[key] = value
    else:
        structured_info = {
            'RPM': 'could not be parsed',
            'NDE/ODE Bearing': 'could not be parsed',
            'DE Bearing': 'could not be parsed',
            'Horsepower': 'could not be parsed',
            'Thermal Class': 'could not be parsed'
        }

    print(f"Structured Information: {structured_info}")

    # Load existing Excel file
    file_path = 'INGD-MDLZ-VA-IME-2024.xlsx'
    wb = load_workbook(file_path)
    ws = wb['Gearboxes']

    # Determine the next available merged row
    next_row = None
    for row in range(4, ws.max_row + 1, 2):
        if ws.cell(row=row, column=2).value is None:
            next_row = row
            break

    if next_row is None:
        next_row = ws.max_row + 1

    # Create a new row as a dictionary
    new_row = {
        'Area': area,
        'Functional location': functional_location,
        'Equipment': equipment,
        'Assembly': assembly,
        'Component': component,
        'Component Picture': '',
        'RPM': structured_info.get('RPM', 'could not be parsed'),
        'NDE Bearing': structured_info.get('NDE/ODE Bearing', 'could not be parsed'),
        'DE Bearing': structured_info.get('DE Bearing', 'could not be parsed'),
        'Horsepower': structured_info.get('Horsepower', 'could not be parsed'),
        'Thermal Class': structured_info.get('Thermal Class', 'could not be parsed')
    }

    # Write the new row data into the worksheet, ensuring all required columns are handled
    col_mapping = {
        'Area': 'B',
        'Functional location': 'C',
        'Equipment': 'D',
        'Assembly': 'E',
        'Component': 'F',
        'Component Picture': 'G',
        'RPM': 'M',
        'NDE Bearing': 'V',
        'DE Bearing': 'W',
        'Horsepower': 'K',
        'Thermal Class': 'S'
    }

    for key, value in new_row.items():
        if key == 'Component Picture':
            continue 
        col = col_mapping[key]
        cell = ws[f"{col}{next_row}"]
        cell.value = value

    # Add the component picture
    if component_picture:
        img = OpenpyxlImage(io.BytesIO(component_picture.read()))
        img.width = 300 
        img.height = 300
        img.anchor = f"G{next_row}"
        ws.add_image(img)

    # Save the workbook
    wb.save(file_path)

    flash('Motor information successfully added!', 'success')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
